"""Biznes hisoboti Excel — 4 varaq: Xulosa · Obunachilar · To'lovlar · Kodlar.

Fayl Google Sheets'ga to'g'ridan-to'g'ri yuklanadi (Fayl → Import), format
saqlanadi. Shu sabab murakkab formula ishlatilmaydi — faqat toza jadval.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.admin_report import BusinessReport

HEADER_BG = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
OK_BG = PatternFill("solid", fgColor="E2EFDA")      # tasdiqlangan
WAIT_BG = PatternFill("solid", fgColor="FFF2CC")    # kutilmoqda
BAD_BG = PatternFill("solid", fgColor="FCE4E4")     # rad etilgan

SUBSCRIBER_HEADERS = (
    ("Telegram ID", 14), ("Username", 18), ("Ism", 24), ("Telefon", 16),
    ("Do'konlar", 28), ("Tarif", 10), ("Holat", 12), ("Sinov tugaydi", 18),
    ("To'langan muddat", 18), ("Qolgan kun", 12), ("Promokod", 18),
    ("Ro'yxatdan o'tgan", 18),
)

PAYMENT_HEADERS = (
    ("№", 8), ("Telegram ID", 14), ("Tarif", 10), ("Summa", 14), ("Oy", 6),
    ("Usul", 12), ("Holat", 14), ("Tashqi ID", 24), ("Yaratilgan", 18),
    ("To'langan", 18),
)

PROMO_HEADERS = (
    ("Kod", 14), ("Tarif", 10), ("Kun", 8), ("Ishlatilgan", 12),
    ("Chegara", 10), ("Faol", 8), ("Muddati", 18), ("Yaratgan", 14),
    ("Izoh", 30),
)


def _write_header(sheet: Worksheet, headers: tuple[tuple[str, int], ...]) -> None:
    for col, (title, width) in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.fill = HEADER_BG
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = "A2"


def build_admin_excel(report: BusinessReport, path: str | Path) -> Path:
    """Biznes hisobotini Excel faylga yozadi."""
    wb = Workbook()

    # --- Varaq 1: Xulosa ---
    s = report.summary
    sheet = wb.active
    sheet.title = "Xulosa"
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 22

    rows: list[tuple[str, object]] = [
        ("Hisobot sanasi", report.generated_at.strftime("%Y-%m-%d %H:%M UTC")),
        ("", ""),
        ("FOYDALANUVCHILAR", ""),
        ("Jami ro'yxatdan o'tgan", s.users),
        ("Do'kon ulagan", s.with_shop),
        ("Faol obuna", s.active_subs),
        ("Promokod bilan kirgan", s.promo_granted),
        ("", ""),
        ("TARIF KESIMIDA (faol)", ""),
    ]
    for plan, count in sorted(s.by_plan.items()):
        rows.append((f"  {plan}", count))

    rows += [
        ("", ""),
        ("PUL", ""),
        ("Tasdiqlangan tushum (so'm)", float(s.paid_total)),
        ("  shundan to'lovlar soni", s.paid_count),
        ("Kutilayotgan (tasdiqlanmagan)", float(s.pending_total)),
        ("  shundan to'lovlar soni", s.pending_count),
        ("Rad etilgan", float(s.rejected_total)),
    ]

    for i, (label, value) in enumerate(rows, start=1):
        cell = sheet.cell(row=i, column=1, value=label)
        if label.isupper() and label:
            cell.font = Font(bold=True, size=12)
        sheet.cell(row=i, column=2, value=value)

    # --- Varaq 2: Obunachilar ---
    sheet = wb.create_sheet("Obunachilar")
    _write_header(sheet, SUBSCRIBER_HEADERS)
    for r, row in enumerate(report.subscribers, start=2):
        for c, value in enumerate(
            (
                row.telegram_id, row.username, row.full_name, row.phone,
                row.shops, row.plan, row.status, row.trial_ends,
                row.paid_until, row.days_left, row.promo_codes, row.registered,
            ),
            start=1,
        ):
            sheet.cell(row=r, column=c, value=value)

    # --- Varaq 3: To'lovlar ---
    sheet = wb.create_sheet("To'lovlar")
    _write_header(sheet, PAYMENT_HEADERS)
    for r, pay in enumerate(report.payments, start=2):
        values = (
            pay.payment_id, pay.telegram_id, pay.plan, float(pay.amount),
            pay.months, pay.method, pay.status, pay.external_id,
            pay.created, pay.paid_at,
        )
        for c, value in enumerate(values, start=1):
            sheet.cell(row=r, column=c, value=value)

        # Holat rangi: tasdiqlangan / kutilmoqda / rad etilgan
        fill = {"paid": OK_BG, "pending": WAIT_BG}.get(pay.status, BAD_BG)
        for c in range(1, len(PAYMENT_HEADERS) + 1):
            sheet.cell(row=r, column=c).fill = fill

    # --- Varaq 4: Promokodlar ---
    sheet = wb.create_sheet("Promokodlar")
    _write_header(sheet, PROMO_HEADERS)
    for r, promo in enumerate(report.promos, start=2):
        values = (
            promo.code, promo.plan, promo.days, promo.used, promo.max_uses,
            "ha" if promo.is_active else "yo'q", promo.expires,
            promo.created_by, promo.note,
        )
        for c, value in enumerate(values, start=1):
            sheet.cell(row=r, column=c, value=value)

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target
