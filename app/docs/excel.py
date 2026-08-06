"""Excel hisobot generatori (SPEC 6.1).

Ustunlar: SKU, shtrix kod, tovar nomi, davr, kutilgan qoldiq, haqiqiy
qoldiq, farq, tannarx, zarar, turi, aniqlangan sana. Pastda jami.

**Shtrix kod majburiy** — usiz seller Uzumga hech narsa isbotlay olmaydi.
Shtrix kodsiz qatorlar ajratib ko'rsatiladi, jimgina yashirilmaydi.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.docs.models import ReportRow

HEADERS = (
    ("SKU", 18),
    ("Shtrix kod", 26),
    ("Tovar nomi", 46),
    ("Davr", 22),
    ("Kutilgan qoldiq", 16),
    ("Haqiqiy qoldiq", 16),
    ("Farq (dona)", 12),
    ("Tannarx (so'm)", 16),
    ("Zarar (so'm)", 16),
    ("Turi", 22),
    ("Aniqlangan sana", 16),
)

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_TOTAL_FONT = Font(bold=True, size=11)
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")  # shtrix kodsiz qatorlar
_MONEY_FMT = "#,##0"
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def build_report(rows: list[ReportRow], output_path: str | Path) -> Path:
    """Excel hisobot yaratadi va yo'lini qaytaradi."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Yo'qotishlar"

    # --- sarlavha qatori ---
    for col, (name, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # --- ma'lumot qatorlari ---
    for idx, row in enumerate(rows, start=2):
        values = (
            row.sku,
            row.barcode or "⚠ YO'Q",
            row.title,
            f"{row.period_from:%d.%m.%Y} — {row.period_to:%d.%m.%Y}",
            row.expected_qty,
            row.actual_qty,
            row.diff_qty,
            float(row.unit_cost),
            float(row.loss_amount),
            row.kind_label,
            row.detected_at,
        )
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.border = _BORDER
            if col in (8, 9):
                cell.number_format = _MONEY_FMT
            elif col == 11:
                cell.number_format = "DD.MM.YYYY"
            elif col in (5, 6, 7):
                cell.alignment = Alignment(horizontal="center")

        # Shtrix kodsiz qator — sariq bilan belgilanadi, chunki bunday
        # tovarga da'vo qilib bo'lmaydi.
        if not row.has_barcode:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=idx, column=col).fill = _WARN_FILL

    # --- jami ---
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="JAMI").font = _TOTAL_FONT
    total_qty = ws.cell(row=total_row, column=7, value=sum(r.diff_qty for r in rows))
    total_amount = ws.cell(
        row=total_row, column=9, value=float(sum(r.loss_amount for r in rows))
    )
    for cell in (total_qty, total_amount):
        cell.font = _TOTAL_FONT
        cell.border = _BORDER
    total_amount.number_format = _MONEY_FMT
    total_qty.alignment = Alignment(horizontal="center")

    # --- shtrix kodsiz qatorlar haqida ogohlantirish ---
    missing = sum(1 for r in rows if not r.has_barcode)
    if missing:
        note = ws.cell(
            row=total_row + 2,
            column=1,
            value=(
                f"⚠ {missing} ta tovarda shtrix kod yo'q (sariq bilan belgilangan). "
                "Bunday tovarlar bo'yicha Uzumga da'vo qilish qiyin — "
                "kartochkada shtrix kodni to'ldiring."
            ),
        )
        note.font = Font(color="9C5700", italic=True)

    wb.save(path)
    return path
