"""Yunit-iqtisodiyot hisoboti — Excel va PDF.

Har SKU: sotuv, tushum, komissiya, logistika, saqlash, sof foyda, marja,
ABC toifasi. Zarar keltiruvchilar ajratib ko'rsatiladi.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.docs.fonts import ensure_fonts
from app.docs.numbers import format_money
from app.docs.pdf import GRID, HEADER_BG, WARN_BG, _styles
from app.services.economics import EconomicsSummary

HEADERS = (
    ("SKU", 16),
    ("Shtrix kod", 20),
    ("Tovar nomi", 42),
    ("ABC", 7),
    ("Sotilgan", 11),
    ("Tushum", 16),
    ("Komissiya", 15),
    ("Logistika", 15),
    ("Saqlash", 14),
    ("Sof foyda", 16),
    ("Marja %", 10),
    ("1 donaga", 14),
)

_LOSS_FILL = PatternFill("solid", fgColor="FFC7CE")  # zarar — qizil


def build_economics_excel(
    summary: EconomicsSummary, output_path: str | Path
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Yunit-iqtisodiyot"

    for col, (name, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for idx, row in enumerate(summary.rows, start=2):
        values = (
            row.sku,
            row.barcode or "—",
            row.title,
            row.abc,
            row.qty_sold,
            float(row.revenue),
            float(row.commission),
            float(row.logistics),
            float(row.storage),
            float(row.profit),
            float(row.margin_pct),
            float(row.profit_per_unit),
        )
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            if col in (6, 7, 8, 9, 10, 12):
                cell.number_format = "#,##0"
            elif col in (4, 5, 11):
                cell.alignment = Alignment(horizontal="center")

        if row.is_loss_making:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=idx, column=col).fill = _LOSS_FILL
        elif row.is_dead_stock:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=idx, column=col).fill = PatternFill(
                    "solid", fgColor="FFF2CC"
                )

    total_row = len(summary.rows) + 2
    ws.cell(row=total_row, column=1, value="JAMI").font = Font(bold=True)
    for col, value in ((6, summary.revenue), (7, summary.commission),
                       (8, summary.logistics), (9, summary.storage),
                       (10, summary.profit)):
        cell = ws.cell(row=total_row, column=col, value=float(value))
        cell.font = Font(bold=True)
        cell.number_format = "#,##0"

    note = ws.cell(
        row=total_row + 2,
        column=1,
        value=(
            "Qizil — zarar keltiruvchi tovarlar. Sariq — sotilmayotgan, "
            "lekin omborda turgan (saqlash puli ketadi)."
        ),
    )
    note.font = Font(italic=True, color="808080")

    wb.save(path)
    return path


def build_economics_pdf(
    summary: EconomicsSummary, output_path: str | Path, *, shop_title: str = ""
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    styles = _styles()
    ensure_fonts()

    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title="Yunit-iqtisodiyot",
    )

    story: list[object] = [Paragraph("Yunit-iqtisodiyot", styles["title"])]
    subtitle = shop_title
    subtitle += f" · {summary.period_from:%d.%m.%Y} — {summary.period_to:%d.%m.%Y}"
    story.append(Paragraph(subtitle, styles["sub"]))

    story.append(
        Paragraph(
            f"Tushum: <b>{format_money(summary.revenue)}</b> so'm · "
            f"Sof foyda: <b>{format_money(summary.profit)}</b> so'm "
            f"(marja {summary.margin_pct}%)",
            styles["cell"],
        )
    )
    story.append(Spacer(1, 4 * mm))

    header = [h[0] for h in HEADERS]
    widths = (18, 24, 52, 10, 14, 24, 22, 22, 20, 24, 14, 20)

    data = [[Paragraph(h, styles["cellHead"]) for h in header]]
    loss_rows: list[int] = []
    dead_rows: list[int] = []
    for index, row in enumerate(summary.rows):
        if row.is_loss_making:
            loss_rows.append(index + 1)
        elif row.is_dead_stock:
            dead_rows.append(index + 1)
        data.append(
            [
                Paragraph(text, styles["cell"])
                for text in (
                    row.sku,
                    row.barcode or "—",
                    row.title,
                    row.abc,
                    str(row.qty_sold),
                    format_money(row.revenue),
                    format_money(row.commission),
                    format_money(row.logistics),
                    format_money(row.storage),
                    format_money(row.profit),
                    f"{row.margin_pct}%",
                    format_money(row.profit_per_unit),
                )
            ]
        )

    table = Table(data, colWidths=[w * mm for w in widths], repeatRows=1)
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
        ("GRID", (0, 0), (-1, -1), 0.4, GRID),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    commands += [
        ("BACKGROUND", (0, i), (-1, i), "#FFC7CE") for i in loss_rows
    ]
    commands += [("BACKGROUND", (0, i), (-1, i), WARN_BG) for i in dead_rows]
    table.setStyle(TableStyle(commands))
    story.append(table)

    if loss_rows or dead_rows:
        story.append(
            Paragraph(
                f"Qizil — zarar keltiruvchi ({len(loss_rows)} ta). "
                f"Sariq — sotilmayotgan, saqlash puli ketadi ({len(dead_rows)} ta).",
                styles["note"],
            )
        )

    doc.build(story)
    return path
