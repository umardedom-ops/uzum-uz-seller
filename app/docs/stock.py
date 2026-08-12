"""Qoldiq hisoboti — Excel va PDF (foydalanuvchi so'rovi bo'yicha).

Sellerga bir ekranda: qayerda nechta qoldi, qancha kunga yetadi, nima
bloklangan. Bu Didox emas — oddiy ombor hisoboti.
"""
from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

from app.docs.fonts import ensure_fonts
from app.docs.numbers import format_money
from app.docs.pdf import GRID, HEADER_BG, WARN_BG, _styles

HEADERS = (
    ("SKU", 16),
    ("Shtrix kod", 22),
    # Sellerning O'Z kodi — ombor xodimi aynan shu bilan ishlaydi,
    # Uzumning SKU raqami ular uchun notanish.
    ("Sotuvchi artikuli", 20),
    ("Tovar nomi", 46),
    ("FBO qoldiq", 12),
    ("FBS qoldiq", 12),
    ("Jami sotilgan", 14),
    ("Kunlik o'rtacha", 14),
    ("Necha kunga yetadi", 18),
    ("Narx (so'm)", 14),
    # Bozor narxi — o'zimiznikini solishtirish uchun
    ("Bozor narxi", 14),
    ("Holat", 20),
)

#: PDF uchun QISQARTIRILGAN ustunlar (nom, kenglik mm).
#: Chop etilib omborga olib boriladi — 12 ustun A4 ga sig'maydi va matn
#: o'qib bo'lmas darajada kichrayadi. Shtrix kod, jami sotilgan va
#: kunlik o'rtacha faqat Excel'da qoladi.
PDF_HEADERS = (
    ("SKU", 20),
    ("Sotuvchi artikuli", 26),
    ("Tovar nomi", 68),
    ("FBO", 16),
    ("FBS", 16),
    ("Necha kunga yetadi", 24),
    ("Narx (so'm)", 24),
    ("Holat", 44),
)


@dataclass(frozen=True, slots=True)
class StockRow:
    sku: str
    barcode: str
    title: str
    fbo_qty: int
    fbs_qty: int
    sold_total: int
    avg_daily_sales: Decimal | None
    price: Decimal | None
    status: str | None
    is_blocked: bool = False
    block_reason: str | None = None
    #: Uzumning O'Z tugash prognozi (kun). Bizning o'rtachadan ustun
    #: turadi — u aksiya va mavsumni ham hisobga oladi.
    forecast_days: int | None = None
    #: Sellerning ichki artikuli (`sellerItemCode` yoki `article`).
    #: Ombor xodimi Uzum SKU raqamini emas, shuni biladi.
    article: str = ""
    #: Bozordagi narx — o'z narxini solishtirish uchun
    market_price: Decimal | None = None

    @property
    def price_gap_pct(self) -> Decimal | None:
        """Bizning narx bozordan necha foiz farq qiladi.

        Musbat — bizniki qimmat, manfiy — arzon. Bozor narxi noma'lum
        bo'lsa None (soxta raqam ko'rsatmaymiz).
        """
        if not self.price or not self.market_price or self.market_price <= 0:
            return None
        diff = (self.price - self.market_price) / self.market_price * 100
        return diff.quantize(Decimal("0.1"))

    @property
    def total_qty(self) -> int:
        return self.fbo_qty + self.fbs_qty

    @property
    def days_left(self) -> int | None:
        """Qoldiq necha kunga yetadi.

        Uzum o'z prognozini bergan bo'lsa — o'sha ustun (aksiya/mavsumni
        hisobga oladi). Bo'lmasa o'rtacha sotuvdan hisoblaymiz. Ikkalasi
        ham yo'q bo'lsa — None (soxta raqam ko'rsatmaymiz).
        """
        if self.forecast_days is not None and self.forecast_days >= 0:
            return self.forecast_days
        if not self.avg_daily_sales or self.avg_daily_sales <= 0:
            return None
        return int(Decimal(self.total_qty) / self.avg_daily_sales)

    @property
    def days_left_label(self) -> str:
        if self.total_qty == 0:
            return "TUGAGAN"
        days = self.days_left
        return f"{days} kun" if days is not None else "—"

    @property
    def status_label(self) -> str:
        if self.is_blocked:
            return f"🚫 BLOKLANGAN: {self.block_reason or 'sabab korsatilmagan'}"
        if self.total_qty == 0:
            return "Tugagan"
        return self.status or "—"

    @property
    def needs_attention(self) -> bool:
        """Diqqat talab qiladi: bloklangan, tugagan yoki 7 kundan kam qolgan."""
        if self.is_blocked or self.total_qty == 0:
            return True
        days = self.days_left
        return days is not None and days <= 7


def build_stock_excel(rows: list[StockRow], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Qoldiqlar"

    for col, (name, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for idx, row in enumerate(rows, start=2):
        values = (
            row.sku,
            row.barcode or "—",
            row.article or "—",
            row.title,
            row.fbo_qty,
            row.fbs_qty,
            row.sold_total,
            float(row.avg_daily_sales) if row.avg_daily_sales else "—",
            row.days_left_label,
            float(row.price) if row.price else "—",
            float(row.market_price) if row.market_price else "—",
            row.status_label,
        )
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            if col in (4, 5, 6, 8):
                cell.alignment = Alignment(horizontal="center")
            elif col == 9:
                cell.number_format = "#,##0"

        if row.needs_attention:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=idx, column=col).fill = PatternFill("solid", fgColor="FFF2CC")

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="JAMI").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(r.fbo_qty for r in rows)).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(r.fbs_qty for r in rows)).font = Font(bold=True)

    wb.save(path)
    return path


def build_stock_pdf(
    rows: list[StockRow], output_path: str | Path, *, shop_title: str = ""
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    styles = _styles()
    ensure_fonts()

    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Qoldiqlar hisoboti",
    )

    story: list[object] = [Paragraph("Qoldiqlar hisoboti", styles["title"])]
    if shop_title:
        story.append(Paragraph(shop_title, styles["sub"]))

    # ❗ PDF Excel'dan KAMROQ ustun oladi. Excel — to'liq ma'lumot uchun,
    # PDF esa chop etib omborga olib boriladi: 12 ustun A4 ga sig'masdi va
    # matn o'qib bo'lmas darajada kichrayardi.
    header = [h[0] for h in PDF_HEADERS]
    widths = [h[1] for h in PDF_HEADERS]

    data = [[Paragraph(h, styles["cellHead"]) for h in header]]
    highlight: list[int] = []
    for index, row in enumerate(rows):
        if row.needs_attention:
            highlight.append(index + 1)
        data.append(
            [
                Paragraph(text, styles["cell"])
                for text in (
                    row.sku,
                    row.article or "—",
                    row.title,
                    str(row.fbo_qty),
                    str(row.fbs_qty),
                    row.days_left_label,
                    format_money(row.price) if row.price else "—",
                    row.status_label,
                )
            ]
        )

    table = Table(data, colWidths=[w * mm for w in widths], repeatRows=1)
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
        ("GRID", (0, 0), (-1, -1), 0.4, GRID),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    commands += [("BACKGROUND", (0, i), (-1, i), WARN_BG) for i in highlight]
    table.setStyle(TableStyle(commands))
    story.append(table)

    attention = sum(1 for r in rows if r.needs_attention)
    if attention:
        story.append(
            Paragraph(
                f"⚠ {attention} ta tovar diqqat talab qiladi: bloklangan, "
                "tugagan yoki 7 kundan kam qolgan (sariq bilan belgilangan).",
                styles["note"],
            )
        )

    doc.build(story)
    return path
