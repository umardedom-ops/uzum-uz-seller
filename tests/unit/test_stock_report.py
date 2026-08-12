"""Qoldiq hisoboti — ustunlar va yangi maydonlar.

❗ Bu testning asosiy vazifasi: sarlavha soni bilan qator soni bir xil
qolishini kafolatlash. HEADERS ga ustun qo'shib, qatorni unutish oson —
o'shanda Excel siljib ketadi yoki PDF yiqiladi (2026-08-11 da aynan shu
tuzoqqa tushildi).
"""
from __future__ import annotations

from decimal import Decimal

from openpyxl import load_workbook

from app.docs.stock import HEADERS, PDF_HEADERS, StockRow, build_stock_excel


def _row(**kw) -> StockRow:
    base = dict(
        sku="763221",
        barcode="1000113258397",
        title="Erkaklar poyabzali",
        fbo_qty=6,
        fbs_qty=2,
        sold_total=40,
        avg_daily_sales=Decimal("1.5"),
        price=Decimal("250000"),
        status="Sotuvda",
    )
    base.update(kw)
    return StockRow(**base)


class TestColumns:
    def test_excel_row_matches_headers(self, tmp_path) -> None:
        path = build_stock_excel([_row()], tmp_path / "q.xlsx")
        sheet = load_workbook(path).active

        assert sheet.max_column == len(HEADERS)
        titles = [sheet.cell(row=1, column=c).value for c in range(1, len(HEADERS) + 1)]
        assert titles == [h[0] for h in HEADERS]

    def test_pdf_is_narrower_than_excel(self) -> None:
        """PDF chop etiladi — kamroq ustun, aks holda A4 ga sig'maydi."""
        assert len(PDF_HEADERS) < len(HEADERS)
        # Landscape A4 ≈ 273 mm foydalanish mumkin
        assert sum(w for _, w in PDF_HEADERS) <= 273


class TestNewFields:
    def test_article_and_market_price_in_excel(self, tmp_path) -> None:
        path = build_stock_excel(
            [_row(article="ART-77", market_price=Decimal("230000"))],
            tmp_path / "q.xlsx",
        )
        sheet = load_workbook(path).active
        values = [sheet.cell(row=2, column=c).value for c in range(1, len(HEADERS) + 1)]

        assert "ART-77" in values
        assert 230000 in values

    def test_missing_values_show_dash(self, tmp_path) -> None:
        """Ma'lumot yo'q bo'lsa — bo'sh emas, «—» (jim qolmaymiz)."""
        path = build_stock_excel([_row(article="", market_price=None)], tmp_path / "q.xlsx")
        sheet = load_workbook(path).active
        values = [sheet.cell(row=2, column=c).value for c in range(1, len(HEADERS) + 1)]
        assert values.count("—") >= 2


class TestPriceGap:
    def test_more_expensive_than_market(self) -> None:
        row = _row(price=Decimal("110000"), market_price=Decimal("100000"))
        assert row.price_gap_pct == Decimal("10.0")

    def test_cheaper_than_market(self) -> None:
        row = _row(price=Decimal("90000"), market_price=Decimal("100000"))
        assert row.price_gap_pct == Decimal("-10.0")

    def test_unknown_market_price_returns_none(self) -> None:
        """Bozor narxi yo'q — soxta raqam chiqarmaymiz."""
        assert _row(market_price=None).price_gap_pct is None
        assert _row(market_price=Decimal("0")).price_gap_pct is None
