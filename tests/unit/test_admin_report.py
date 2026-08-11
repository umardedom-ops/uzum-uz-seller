"""Biznes hisoboti — obunachilar, to'lovlar (tasdiqlangan/yo'q), kodlar."""
from __future__ import annotations

from datetime import timedelta
from decimal import Decimal

from openpyxl import load_workbook

from app.db.base import session_scope, utcnow
from app.db.models import (
    Payment,
    PaymentMethod,
    PaymentStatus,
    Plan,
    PromoCode,
    PromoRedemption,
    Shop,
    Subscription,
    SubscriptionStatus,
    User,
)
from app.docs.admin_report import build_admin_excel
from app.services import admin_report


async def _seed() -> None:
    now = utcnow()
    async with session_scope() as s:
        u1 = User(telegram_id=1001, username="seller_a", full_name="Ali")
        u2 = User(telegram_id=1002, username="seller_b", full_name="Vali")
        s.add_all([u1, u2])
        await s.flush()

        s.add(Shop(user_id=u1.id, uzum_shop_id="7973", title="AZIKO", is_active=True))

        s.add(
            Subscription(
                user_id=u1.id,
                plan=Plan.PRO,
                status=SubscriptionStatus.ACTIVE,
                paid_until=now + timedelta(days=30),
            )
        )
        s.add(
            Subscription(
                user_id=u2.id,
                plan=Plan.BASIC,
                status=SubscriptionStatus.TRIAL,
                trial_ends_at=now + timedelta(days=2),
            )
        )

        # Tasdiqlangan va tasdiqlanmagan to'lov
        s.add(
            Payment(
                user_id=u1.id, plan=Plan.PRO, amount=Decimal("299000"), months=1,
                method=PaymentMethod.CLICK, status=PaymentStatus.PAID, paid_at=now,
            )
        )
        s.add(
            Payment(
                user_id=u2.id, plan=Plan.BASIC, amount=Decimal("149000"), months=1,
                method=PaymentMethod.MANUAL, status=PaymentStatus.PENDING,
            )
        )

        promo = PromoCode(code="TESTKOD1", plan=Plan.PRO, days=90, max_uses=10)
        s.add(promo)
        await s.flush()
        s.add(PromoRedemption(promo_id=promo.id, user_id=u2.id))
        promo.used_count = 1


class TestCollect:
    async def test_counts_and_money(self) -> None:
        await _seed()
        rep = await admin_report.collect()

        assert rep.summary.users == 2
        assert rep.summary.with_shop == 1
        assert rep.summary.active_subs == 2
        # Tasdiqlangan va kutilayotgan pul alohida sanaladi
        assert rep.summary.paid_total == Decimal("299000")
        assert rep.summary.pending_total == Decimal("149000")
        assert rep.summary.paid_count == 1
        assert rep.summary.pending_count == 1

    async def test_promo_linked_to_user(self) -> None:
        """Kim qaysi kod orqali kirgani ko'rinishi kerak."""
        await _seed()
        rep = await admin_report.collect()

        by_tg = {r.telegram_id: r for r in rep.subscribers}
        assert "TESTKOD1" in by_tg[1002].promo_codes
        assert by_tg[1001].promo_codes == ""
        assert rep.summary.promo_granted == 1

    async def test_shop_shown(self) -> None:
        await _seed()
        rep = await admin_report.collect()
        by_tg = {r.telegram_id: r for r in rep.subscribers}
        assert "AZIKO" in by_tg[1001].shops


class TestManagementView:
    """Boshqaruv savollari: kim Pro, kim sinovda, bu oy nima bo'ldi."""

    async def test_pro_and_trial_are_separated(self) -> None:
        """Sinovdagi odam `effective_plan` da BASIC ko'rinadi, lekin u
        to'lovchi EMAS — hisobda aralashib ketmasligi kerak."""
        await _seed()
        rep = await admin_report.collect()

        assert rep.summary.pro_paid == 1      # u1 — Pro sotib olgan
        assert rep.summary.basic_paid == 0
        assert rep.summary.on_trial == 1      # u2 — sinovda
        assert rep.summary.expired == 0

    async def test_this_month_numbers(self) -> None:
        await _seed()
        rep = await admin_report.collect()

        # Ikkalasi ham hozir yaratildi — ya'ni shu oyda
        assert rep.summary.joined_this_month == 2
        assert rep.summary.payers_this_month == 1
        assert rep.summary.paid_this_month == Decimal("299000")
        assert rep.summary.month_label  # "YYYY-MM"

    async def test_row_marks_month_and_source(self) -> None:
        await _seed()
        rep = await admin_report.collect()
        by_tg = {r.telegram_id: r for r in rep.subscribers}

        assert by_tg[1001].joined_this_month
        assert by_tg[1001].paid_this_month == Decimal("299000")
        assert by_tg[1001].source == "to'lov"
        # u2 promokod bilan kirgan
        assert by_tg[1002].source == "promokod"

    async def test_average_payment(self) -> None:
        await _seed()
        rep = await admin_report.collect()
        assert rep.summary.paid_count == 1
        assert rep.summary.paid_total == Decimal("299000")


class TestExcel:
    async def test_four_sheets_with_data(self, tmp_path) -> None:
        await _seed()
        rep = await admin_report.collect()

        path = build_admin_excel(rep, tmp_path / "hisobot.xlsx")
        wb = load_workbook(path)

        assert wb.sheetnames == ["Xulosa", "Obunachilar", "To'lovlar", "Promokodlar"]
        # Sarlavha + 2 obunachi
        assert wb["Obunachilar"].max_row == 3
        assert wb["To'lovlar"].max_row == 3
        assert wb["Promokodlar"].max_row == 2

    async def test_empty_db_does_not_crash(self, tmp_path) -> None:
        """Ma'lumot yo'q bo'lsa ham fayl chiqadi — jim yiqilmaymiz."""
        rep = await admin_report.collect()
        path = build_admin_excel(rep, tmp_path / "bosh.xlsx")
        assert path.exists()
